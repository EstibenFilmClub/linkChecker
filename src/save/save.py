import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta
import os


class ExcelRecorder:
    def __init__(self, filename="video_info.xlsx"):
        self.filename = os.path.join("src", "data", filename)  # Guardar en la carpeta src/data
        self.initialize_excel()

    def initialize_excel(self):
        # Crear un archivo Excel o añadir encabezados si no existen
        if not os.path.exists(os.path.dirname(self.filename)):
            os.makedirs(os.path.dirname(self.filename))  # Crear las carpetas si no existen

        try:
            workbook = openpyxl.load_workbook(self.filename)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            # Añadir encabezados
            sheet.append(["URL", "Título","Fecha Última Comprobacion", "Estado" ])

        workbook.save(self.filename)

    def add_video_info(self, url, title, estado):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook.active

        # Buscar si la URL ya existe
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == url:
                last_time = datetime.strptime(sheet.cell(
                    row=row, column=3).value, "%Y-%m-%d %H:%M:%S")
                if datetime.now() - last_time > timedelta(weeks=1):
                    # Actualizar registro
                    sheet.cell(row=row, column=2).value = title
                    sheet.cell(row=row, column=3).value = datetime.now().strftime(
                        "%Y-%m-%d %H:%M:%S")
                    sheet.cell(row=row, column=4).value = estado  # Actualizar el estado
                    workbook.save(self.filename)
                return

        # Añadir la información en una nueva fila si la URL no existe
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([url, title, current_time, estado])  # Incluir el estado

        # Guardar el archivo
        workbook.save(self.filename)

